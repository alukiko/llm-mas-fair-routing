import os
from dataclasses import dataclass
from pathlib import Path
from typing import Annotated, List, Optional

from fastmcp import Context, FastMCP
from markitdown import MarkItDown
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from pydantic import Field

from automas.mcp.servers.content_utils import create_temp_directory, hash_file, truncate_text
from automas.mcp.servers.document.image_utils import (
    format_image_section,
    get_cached_images,
    save_cached_images,
)

CACHE_DIR = Path.home() / ".automas" / "xlsx_cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

md = MarkItDown()

xlsx_server = FastMCP("xlsx-reader")


@dataclass
class ExtractedXLSXImage:
    path: str
    sheet_name: str
    cell: str
    width: int
    height: int


async def extract_images_from_xlsx(xlsx_path: str, ctx: Context) -> List[ExtractedXLSXImage]:
    xlsx_hash = hash_file(xlsx_path)

    cached = await get_cached_images(CACHE_DIR, xlsx_hash, ExtractedXLSXImage)
    if cached is not None:
        await ctx.info(f"Using cached images ({len(cached)} images)")
        return cached

    output_dir = create_temp_directory("xlsx_images")

    images = []
    wb = None

    try:
        wb = load_workbook(xlsx_path)
        total_sheets = len(wb.sheetnames)

        await ctx.info(f"Extracting images from {total_sheets} sheets")

        for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            if total_sheets > 1:
                await ctx.report_progress(progress=sheet_idx, total=total_sheets)

            sheet = wb[sheet_name]

            try:
                image_loader = SheetImageLoader(sheet)

                for img_cell in image_loader._images.keys():
                    try:
                        image = image_loader.get(img_cell)

                        filename = f"sheet{sheet_idx}_{img_cell}.png"
                        image_path = output_dir / filename

                        image.save(image_path)

                        images.append(
                            ExtractedXLSXImage(
                                path=str(image_path),
                                sheet_name=sheet_name,
                                cell=img_cell,
                                width=image.width,
                                height=image.height,
                            )
                        )
                    except Exception:
                        pass

            except Exception:
                pass

    except Exception as e:
        raise Exception(f"Error extracting images from XLSX: {e}")
    finally:
        if wb is not None:
            wb.close()

    await save_cached_images(CACHE_DIR, xlsx_hash, images)

    if images:
        await ctx.info(f"Extracted {len(images)} images from XLSX")

    return images


@xlsx_server.tool
async def read_xlsx_xls(
    file_path: Annotated[str, Field(description="Path to the XLSX or XLS file")],
    ctx: Context,
    max_lines: Annotated[
        Optional[int],
        Field(description="Maximum number of text lines to return. If None, returns all lines"),
    ] = 1000,
) -> str:
    """
    Read an XLSX or XLS file and return its text content.

    Args:
        file_path: Path to the XLSX or XLS file (Excel spreadsheet)
        max_lines: Maximum number of text lines to return (default: 1000)

    Returns:
        Text content of the XLSX or XLS file
    """
    try:
        expanded_path = os.path.expanduser(file_path)
        file_name = os.path.basename(file_path)

        await ctx.info(f"Reading XLSX: {file_name}")

        text_content = md.convert(expanded_path).text_content
        text_content = truncate_text(text_content, max_lines)

        images = await extract_images_from_xlsx(expanded_path, ctx)

        text_content += format_image_section(
            images, [("path", "Path"), ("sheet_name", "Sheet"), ("cell", "Cell")]
        )

        return text_content

    except Exception as e:
        await ctx.error(f"Failed to read XLSX/XLS: {e}")
        return f"Error reading XLSX/XLS: {e}"
